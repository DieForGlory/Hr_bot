# -*- coding: utf-8 -*-
"""Дополняет справочник сотрудников (employees_seed.json) данными по отпускам
из кадрового Excel «ОТПУСКА» (GHPG.xlsx): дата приёма и уже потраченные дни.

Использование:
    python scripts/merge_vacation_xlsx.py <GHPG.xlsx> [data/employees_seed.json] [--entity "GOLDEN HOUSE PROPERTY GROUP"]

Из Excel берутся колонки:
    A  Юрлицо                     -> фильтр (см. ниже)
    C  Сотрудник                  -> сопоставление по ФИО
    F  Дата приема                -> hire_date
    S  Потрачено дней рабочих     -> used_work_days
    T  Потрачено дней календарных -> used_calendar_days

ВАЖНО: в Excel один и тот же человек встречается несколько раз — по строке на
каждое юрлицо холдинга, где он оформлен, и у каждой записи СВОЯ дата приёма и
свои потраченные дни. Бот обслуживает одно юрлицо, поэтому берём только строки
нужной компании (--entity, по умолчанию GOLDEN HOUSE PROPERTY GROUP).

Остаток (U/V) НЕ импортируется — он считается ботом на лету
(bot/utils/vacation_balance.py) из даты приёма и потраченных дней.

ВНИМАНИЕ: и Excel, и результат содержат персональные данные — в репозитории
не хранятся, файл живёт в томе ./data рядом с БД.
"""
import json
import os
import re
import sys
import difflib
import unicodedata
from datetime import date

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from bot.utils.vacation_balance import compute_vacation_balance, parse_hire_date  # noqa: E402

# Колонки (1-based) листа «ОТПУСКА»
COL_ENTITY = 1    # A Юрлицо
COL_POSITION = 2  # B Должность
COL_NAME = 3      # C Сотрудник
COL_HIRE = 6      # F Дата приема
COL_ACCRUED_WORK = 16      # P Дни рабочие (начислено)
COL_ACCRUED_CALENDAR = 18  # R Дни календарные (начислено)
COL_USED_WORK = 19      # S Потрачено дней рабочих
COL_USED_CALENDAR = 20  # T Потрачено дней календарных
FIRST_DATA_ROW = 3

DEFAULT_ENTITY = "GOLDEN HOUSE PROPERTY GROUP"

_APOSTROPHES = "'`‘’ʻʼ´"


def norm_entity(s: str) -> str:
    """Юрлицо -> ключ для сравнения: только буквы/цифры, верхний регистр.
    'ООО \"GOLDEN HOUSE PROPERTY GROUP\" ' -> 'ОООGOLDENHOUSEPROPERTYGROUP'."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return re.sub(r"[^\w]", "", s, flags=re.UNICODE).upper()


def norm_name(s: str) -> str:
    """Нормализация ФИО для сопоставления: без апострофов, регистра и лишних пробелов."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    for ch in _APOSTROPHES:
        s = s.replace(ch, "")
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s).strip().upper()
    # ФИО, продублированное в одной ячейке: "X Y Z X Y Z" -> "X Y Z"
    words = s.split()
    if len(words) % 2 == 0:
        half = len(words) // 2
        if words[:half] == words[half:]:
            s = " ".join(words[:half])
    return s


def _num(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _hire_date(value) -> str | None:
    """Дата приёма -> строка 'dd.mm.yyyy' (в Excel встречается и текстом, и датой)."""
    if value is None or value == "":
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")
    s = str(value).strip()
    return s or None


def load_vacation_rows(path: str, entity_key: str) -> tuple[list[dict], date]:
    """Строки нужного юрлица + дата, на которую посчитан Excel (W1 = TODAY())."""
    ws = openpyxl.load_workbook(path, data_only=True).worksheets[0]

    w1 = ws.cell(1, 23).value  # W1
    as_of = w1.date() if hasattr(w1, "date") else date.today()

    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        name = ws.cell(r, COL_NAME).value
        if not name or not str(name).strip():
            continue
        if entity_key and entity_key not in norm_entity(ws.cell(r, COL_ENTITY).value):
            continue
        rows.append({
            "row": r,
            "entity": ws.cell(r, COL_ENTITY).value,
            "position": ws.cell(r, COL_POSITION).value,
            "name_raw": str(name).strip(),
            "name_norm": norm_name(name),
            "hire_date": _hire_date(ws.cell(r, COL_HIRE).value),
            "accrued_work": _num(ws.cell(r, COL_ACCRUED_WORK).value),
            "accrued_calendar": _num(ws.cell(r, COL_ACCRUED_CALENDAR).value),
            "used_work_days": _num(ws.cell(r, COL_USED_WORK).value),
            "used_calendar_days": _num(ws.cell(r, COL_USED_CALENDAR).value),
        })
    return rows, as_of


def main():
    argv = sys.argv[1:]
    entity = DEFAULT_ENTITY
    if "--entity" in argv:
        i = argv.index("--entity")
        entity = argv[i + 1]
        del argv[i:i + 2]

    if not argv:
        print(__doc__)
        sys.exit(1)
    xlsx_path = argv[0]
    seed_path = argv[1] if len(argv) > 1 else os.path.join("data", "employees_seed.json")

    for p in (xlsx_path, seed_path):
        if not os.path.exists(p):
            print(f"Файл не найден: {p}")
            sys.exit(1)

    with open(seed_path, encoding="utf-8") as f:
        seed = json.load(f)

    entity_key = norm_entity(entity)
    vac_rows, as_of = load_vacation_rows(xlsx_path, entity_key)
    print(f"Юрлицо: {entity!r} -> строк в Excel: {len(vac_rows)}")
    print(f"Excel посчитан на дату (W1): {as_of}\n")
    if not vac_rows:
        print("Ни одной строки по этому юрлицу — проверьте --entity")
        sys.exit(1)

    # индекс по нормализованному ФИО (дубли внутри Excel собираем в список)
    by_norm: dict[str, list[dict]] = {}
    for v in vac_rows:
        by_norm.setdefault(v["name_norm"], []).append(v)

    norm_keys = list(by_norm.keys())
    matched = 0
    fuzzy_used = []
    unmatched_seed = []
    ambiguous = []
    frozen = []

    for rec in seed:
        key = norm_name(rec["full_name"])
        hit = by_norm.get(key)
        how = "exact"
        if not hit:
            cand = difflib.get_close_matches(key, norm_keys, n=1, cutoff=0.90)
            if cand:
                hit = by_norm[cand[0]]
                how = "fuzzy"
                fuzzy_used.append((rec["full_name"], hit[0]["name_raw"]))
        if not hit:
            unmatched_seed.append(rec["full_name"])
            rec.setdefault("hire_date", None)
            rec.setdefault("used_work_days", 0)
            rec.setdefault("used_calendar_days", 0)
            rec.setdefault("accrued_work_override", None)
            rec.setdefault("accrued_calendar_override", None)
            continue

        if len(hit) > 1:
            ambiguous.append((rec["full_name"], [h["name_raw"] for h in hit]))
        v = hit[0]
        rec["hire_date"] = v["hire_date"]
        rec["used_work_days"] = v["used_work_days"]
        rec["used_calendar_days"] = v["used_calendar_days"]
        matched += 1

        # Нестандартное начисление? Считаем штатным алгоритмом на дату Excel и
        # сравниваем с колонками P/R. Расхождение = HR вручную обрезал период
        # (декрет и т.п.) -> фиксируем начисление снимком, иначе бот насчитает лишнее.
        hire = parse_hire_date(v["hire_date"])
        if hire is None:
            rec["accrued_work_override"] = v["accrued_work"]
            rec["accrued_calendar_override"] = v["accrued_calendar"]
            frozen.append((rec["full_name"], "нет даты приёма"))
            continue

        std = compute_vacation_balance(hire, today=as_of)
        diff_work = abs(std.accrued_work - v["accrued_work"]) > 0.01
        diff_cal = abs(std.accrued_calendar - v["accrued_calendar"]) > 0.01
        if diff_work or diff_cal:
            rec["accrued_work_override"] = v["accrued_work"]
            rec["accrued_calendar_override"] = v["accrued_calendar"]
            frozen.append((
                rec["full_name"],
                f"начислено раб. {std.accrued_work}->{v['accrued_work']}, "
                f"кал. {std.accrued_calendar}->{v['accrued_calendar']} ({rec.get('work_state') or '-'})"
            ))
        else:
            rec["accrued_work_override"] = None
            rec["accrued_calendar_override"] = None

    matched_norms = {norm_name(r["full_name"]) for r in seed if r.get("hire_date")}
    only_in_excel = [v["name_raw"] for v in vac_rows if v["name_norm"] not in matched_norms]

    with open(seed_path, "w", encoding="utf-8") as f:
        json.dump(seed, f, ensure_ascii=False, indent=1)

    print(f"Записей в справочнике: {len(seed)}")
    print(f"Строк по юрлицу: {len(vac_rows)}")
    print(f"Сопоставлено (есть дата приёма): {matched}")
    print(f"  из них fuzzy: {len(fuzzy_used)}")
    print(f"  из них с «замороженным» начислением (снимок): {len(frozen)}")
    print(f"Без данных по отпускам: {len(unmatched_seed)}")
    print(f"Есть в Excel по этому юрлицу, но нет в справочнике: {len(only_in_excel)}")

    if frozen:
        print(f"\n--- НАЧИСЛЕНИЕ ЗАФИКСИРОВАНО СНИМКОМ ({len(frozen)}) ---")
        for n, why in frozen:
            print(f"  {n!r}: {why}")

    if fuzzy_used:
        print("\n--- FUZZY-сопоставления (проверьте) ---")
        for a, b in fuzzy_used:
            print(f"  {a!r} <- {b!r}")
    if ambiguous:
        print("\n--- ДУБЛИ ВНУТРИ ОДНОГО ЮРЛИЦА (взята первая строка — проверьте!) ---")
        for a, b in ambiguous:
            print(f"  {a!r} -> {b}")
    if only_in_excel:
        print(f"\n--- ЕСТЬ В EXCEL ПО ЮРЛИЦУ, НЕТ В СПРАВОЧНИКЕ ({len(only_in_excel)}) ---")
        for n in only_in_excel:
            print(f"  {n!r}")
    if unmatched_seed:
        print(f"\n--- БЕЗ ДАННЫХ ПО ОТПУСКАМ ({len(unmatched_seed)}) ---")
        for n in unmatched_seed:
            print(f"  {n!r}")

    print(f"\nЗаписано: {seed_path}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
